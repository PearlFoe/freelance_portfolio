import requests
import openpyxl

from time import process_time
from bs4 import BeautifulSoup as bs
from fake_useragent import UserAgent

main_url = 'https://avtoshkoli.ru'
cities_urls = []
all_schools_inf = []

def get_cities_urls(url):
	response = requests.get(url)
	soup = bs(response.text, 'lxml')
	root = soup.find('div', {'class':'main_cities_bg'}).find_all('li')

	for i in root:
		if len(i.get_text()) > 1:
			cities_urls.append(url + i.a.attrs['href'])

def get_school_inf(city_url):
	school_inf = []

	response = requests.get(city_url)
	soup = bs(response.text, 'lxml')
	schools = soup.find('div', {'class':'firms_list_items'}).find_all('a')

	for school in schools:
		school_inf.append(school.find('div', {'class':'firm_thumb_title'}).get_text())
		school_inf.append(school.find('div', {'class':'firm_thumb_contacts_unit firm_thumb_contacts_address'}).get_text())

		try:
			school_inf.append(school.find('div', {'class':'firm_thumb_contacts_unit firm_thumb_contacts_phone'}).get_text())
		except Exception:
			school_inf.append('-')

		school_inf.append(school.find('div', {'class':'firm_thumb_rating'}).b.get_text())

		try:
			school_inf.append(school.find('div', {'class':'firm_thumb_price'}).b.get_text())
		except Exception:
			school_inf.append(school.find('div', {'class':'firm_thumb_price'}).i.get_text())

		school_inf.append(city_url[:-1] + school.attrs['href'])

		all_schools_inf.append(school_inf)
		school_inf = []

def main():
	print('Scrapping cityies urls\n')
	get_cities_urls(main_url)

	print('Scrapping schools information in each city\n')
	for i in cities_urls:
		get_school_inf(i)

	print('Saving all scrapped information in database\n')
	wb = openpyxl.Workbook()
	sheet = wb['Sheet']

	for row in range(1, len(all_schools_inf) + 2):
		for col in range(1, 7):
			if row == 1:
				head = ['Название','Адресс','Телефон','Рейтинг','Цена(категория В)','Ссылка на страницу']
				value = head[col - 1]
				cell = sheet.cell(row = row, column = col)
				cell.value = value
			else:
				value = all_schools_inf[row - 2][col - 1]
				cell = sheet.cell(row = row, column = col)
				cell.value = value

	wb.save('driving_schools.xlsx')
	print('Done')

if __name__ == "__main__":
	main()