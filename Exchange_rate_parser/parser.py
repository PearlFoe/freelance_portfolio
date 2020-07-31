import requests
from bs4 import BeautifulSoup as bs
import openpyxl
import numpy as np
from fake_useragent import UserAgent
import time

url = 'https://www.bestchange.ru/list.html'
proxies = {'https': '93.171.164.251:8080'}

def parse_head(url, proxies):
	html = requests.get(url, proxies, headers={'User-Agent': UserAgent().chrome}).text
	soup = bs(html, 'lxml')
	head = soup.find('table', {'id':'content_table'}).thead.get_text()
	
	head_arr = 	head.split('\n')
	
	for i in head_arr:
		if len(i) < 3:
			head_arr.remove(i)

	head_arr.remove(head_arr[-1])
	head_arr.remove(head_arr[-1])

	return head_arr

def parse_body(url, proxies):
	html = requests.get(url, proxies).text
	soup = bs(html, 'lxml')
	body = soup.find('tbody').find_all('td')

	body_arr = []

	for i in body:
		if len(i.get_text()) == 0:
			body.remove(i)

	for i in body:
		body_arr.append(i.get_text())

	try:
		while True:
			index = body_arr.index('')
			body_arr[index] = ' '
			body_arr[index - 1] = ' '
			body_arr[index - 2] = ' '
			body_arr[index - 3] = ' '
	except Exception:
		while True:
			try:
				index = body_arr.index(' ')
				body_arr.remove(body_arr[index])
			except Exception:
				break

	return body_arr

def main():
	head = parse_head(url, proxies)
	time.sleep(2)
	body = parse_body(url, proxies)
	
	wb = openpyxl.Workbook()
	sheet = wb['Sheet']
	
	cell_counter = 0

	for row in range(1, 451):
		for col in range(1, 9):
			if row == 1:
				value = head[col - 1]
				cell = sheet.cell(row = row, column = col)
				cell.value = value
			else:
				value = body[cell_counter]
				cell = sheet.cell(row = row, column = col)
				cell.value = value
				cell_counter += 1

	wb.save('example.xlsx')


if __name__ == "__main__":
	main()